import io
import re
import pdfplumber
from typing import Optional
from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from app.config import BANCO_AUTENTICACAO
from app.database import obter_conexao
from app.utils import corrigir_encoding
from .pdf_bancos import (
    pdf_banco_do_brasil,
    pdf_bradesco,
    pdf_vp,
    pdf_caixa,
    pdf_sicoob,
    pdf_safra,
    pdf_sicredi,
    pdf_c6,
    pdf_inter,
    pdf_cora,
    pdf_stone,
)

router = APIRouter(prefix="/NashBancoConsultoria/conversor", tags=["Conversor de Extratos"])

# ==============================================================================
# ENRIQUECIMENTO E APLICAÇÃO DE REGRAS
# ==============================================================================

def corrigir_tags_ofx_abertas(conteudo_bytes: bytes) -> str:
    """
    Lê o conteúdo em bytes de um arquivo OFX, detecta tags que foram abertas 
    e adiciona a tag de fechamento correspondente de forma segura.
    """
    try:
        try:
            texto = conteudo_bytes.decode('utf-8')
        except UnicodeDecodeError:
            texto = conteudo_bytes.decode('latin-1')
 
        linhas = texto.splitlines()
        linhas_corrigidas = []
 
        # Regex para encontrar tags no formato <TAG>valor (sem fechamento na mesma linha)
        padrao_tag = re.compile(r'^\s*<([A-Za-z0-9_]+)>([^<\r\n]+)$')
 
        for linha in linhas:
            match = padrao_tag.match(linha)
            if match:
                nome_tag = match.group(1)
                
                # Evita mexer em tags que já possuem fechamento ou tags de cabeçalho
                if not f"</{nome_tag}>" in linha and not nome_tag.startswith("OFX"):
                    linha_formatada = f"<{nome_tag}>{match.group(2).strip()}</{nome_tag}>"
                    linhas_corrigidas.append(linha_formatada)
                    continue
            
            linhas_corrigidas.append(linha)
 
        return "\n".join(linhas_corrigidas)
 
    except Exception as e:
        print(f"Aviso: Erro ao tentar corrigir tags do OFX: {e}")
        return conteudo_bytes.decode('utf-8', errors='ignore')

def identificar_banco_no_pdf(conteudo_bytes: bytes) -> str:
    with pdfplumber.open(io.BytesIO(conteudo_bytes)) as pdf:
        if not pdf.pages:
            return ""
        texto_pagina1 = (pdf.pages[0].extract_text() or "").upper()
        texto_completo = "\n".join(
            (page.extract_text() or "").upper() for page in pdf.pages)
        
        if "V & P DISTRIBUIDORA" in texto_completo and "EXTRATO PARA" in texto_completo:
            return "V&P"

        if "BANCO DO BRASIL" in texto_pagina1:
            return "BANCO DO BRASIL"
        if "LOTE" in texto_pagina1 and "DOCUMENTO" in texto_pagina1 and "HISTÓRICO" in texto_pagina1:
            return "BANCO DO BRASIL"
        if "BRADESCO" in texto_pagina1:
            return "BRADESCO"
        if "NET EMPRESA" in texto_pagina1 or "TOTAL DISPONÍVEL (R$)" in texto_pagina1:
            return "BRADESCO"
        if ("LANÇAMENTO" in texto_pagina1 and "DCTO." in texto_pagina1 and "CRÉDITO (R$)" in texto_pagina1):
            return "BRADESCO"
        if "CAIXA" in texto_pagina1:
            return "CAIXA"
        if "SICOOB" in texto_pagina1:
            return "SICOOB"
        if "SAFRA" in texto_pagina1:
            return "BANCO SAFRA"
        if "SICREDI" in texto_completo:
            return "SICREDI"
        if "C6 BANK" in texto_completo or "BANCO C6" in texto_completo:
            return "BANCO C6"
        if "BANCO INTER" in texto_pagina1:
            return "BANCO INTER"
        if "CORA SCFI" in texto_pagina1 or ("CORA" in texto_pagina1 and "PAGAMENTOS S.A" in texto_pagina1):
            return "BANCO CORA"
        if "STONE" in texto_pagina1:
            return "BANCO STONE"
            
    return ""

def identificar_fornecedor(descricao: str, banco_val: str, tipo: str = "", conexao = None) -> str:
    desc_upper = (descricao or "").upper().strip()
    tipo_upper = (tipo or "").upper().strip()

    if conexao:
        try:
            cursor = conexao.cursor()
            query = """
                SELECT 
                    fr.termoDescricao,
                    fr.termoTipo,
                    f.nome AS nomeFornecedor
                FROM dbo.FornecedorRegras fr
                INNER JOIN dbo.Fornecedor f ON fr.fornecedorId = f.id
                WHERE f.status = 1
                ORDER BY 
                    fr.prioridade DESC,           -- 1º Critério: Maior prioridade definida
                    LEN(fr.termoDescricao) DESC,  -- 2º Critério: Termos mais específicos (desempate)
                    fr.id DESC                    -- 3º Critério: Mais recentes primeiro (desempate)
            """
            cursor.execute(query)
            regras = cursor.fetchall()

            for t_desc, t_tipo, nome_fornecedor in regras:
                match_desc = True if not t_desc else (t_desc.upper() in desc_upper)
                match_tipo = True if not t_tipo else (t_tipo.upper() == tipo_upper)

                if match_desc and match_tipo:
                    fornecedor_upper = str(nome_fornecedor).strip().upper()
                    if fornecedor_upper in ["BANCO", "(NOME DO BANCO)", "[NOME DO BANCO]"]:
                        return banco_val
                    return nome_fornecedor
        except Exception as e:
            print(f"Aviso: Erro ao buscar regras de fornecedores: {e}")

    return ""

def aplicar_plano_conta(linhas_extrato: list, contratante_id: Optional[int] = None) -> list:
    """
    Busca o nome do contratante e aplica as regras de DE/PARA cadastradas na dbo.PlanoDePara,
    preenchendo: contratante, planoConta, grupoConta e edre.
    """
    nome_contratante = ""
    regras = []

    if contratante_id:
        conexao = obter_conexao(BANCO_AUTENTICACAO)
        cursor = conexao.cursor()

        # 1. Busca o nome do contratante
        try:
            cursor.execute("SELECT nome FROM dbo.Contratante WHERE id = ?", [contratante_id])
            row_contratante = cursor.fetchone()
            if row_contratante:
                nome_contratante = row_contratante[0]
        except Exception as e:
            print(f"Aviso: Erro ao buscar nome do contratante: {e}")

        # 2. Busca as regras ativas associando com dbo.Fornecedor (se houver fornecedorId)
        try:
            query_regras = """
                SELECT 
                    p.termoDescricao,
                    f.nome AS termoFornecedor,
                    p.termoTipo,
                    pc.planoConta,
                    pc.grupoConta,
                    pc.edre
                FROM dbo.PlanoDePara p
                INNER JOIN dbo.PlanoContas pc ON p.planoContaId = pc.id
                LEFT JOIN dbo.Fornecedor f ON p.fornecedorId = f.id
                WHERE p.contratanteId IS NULL OR p.contratanteId = ?
            """
            cursor.execute(query_regras, [contratante_id])
            regras = cursor.fetchall()
        except Exception as e:
            print(f"Aviso: Erro ao buscar regras de PlanoDePara: {e}")
        finally:
            conexao.close()

    # 3. Preenche as linhas do extrato com o contratante e as categorias mapeadas
    for linha in linhas_extrato:
        linha["contratante"] = nome_contratante

        desc = (linha.get("descricao") or "").upper().strip()
        forn = (
            linha.get("fornecedor")
            or linha.get("fornecedores")
            or ""
        ).upper().strip()
        
        tipo_linha = (linha.get("tipo") or "").upper().strip()

        for t_desc, t_forn, t_tipo, p_conta, g_conta, edre in regras:
            match_desc = True if not t_desc else (t_desc.upper() in desc)
            match_forn = True if not t_forn else (t_forn.upper() in forn)
            match_tipo = True if not t_tipo else (t_tipo.upper() == tipo_linha)

            if match_desc and match_forn and match_tipo:
                linha["planoConta"] = p_conta or ""
                linha["grupoConta"] = g_conta or ""
                linha["edre"] = edre or ""
                break

    return linhas_extrato

# ==============================================================================
# PARSER OFX E ROTEADOR
# ==============================================================================
def processar_ofx(conteudo_bytes_ou_texto, conexao) -> list:
    """
    Processa o conteúdo do arquivo OFX suportando tags abertas,
    validando o banco cadastrado e vinculando Unidade e Contratante.
    """
    cursor = conexao.cursor()

    # 1. Carrega o mapeamento de Bancos da tabela dbo.Banco (chave: código sem zeros à esquerda)
    cursor.execute("SELECT codigo, nome FROM dbo.Banco WHERE codigo IS NOT NULL")
    mapa_bancos_db = {}
    for row in cursor.fetchall():
        cod_db, nome_db = row[0], row[1]
        if cod_db:
            mapa_bancos_db[str(cod_db).strip().lstrip("0")] = nome_db.strip()

    # 2. Tratamento inicial e decodificação do arquivo
    if isinstance(conteudo_bytes_ou_texto, bytes):
        bytes_corrigidos = conteudo_bytes_ou_texto
    else:
        bytes_corrigidos = str(conteudo_bytes_ou_texto).encode("latin-1", errors="ignore")

    # Corrige tags OFX abertas antes de processar as expressões regulares
    if 'corrigir_tags_ofx_abertas' in globals():
        conteudo_corrigido_bytes = corrigir_tags_ofx_abertas(bytes_corrigidos)
    else:
        conteudo_corrigido_bytes = bytes_corrigidos

    if isinstance(conteudo_corrigido_bytes, bytes):
        try:
            conteudo_texto = conteudo_corrigido_bytes.decode('utf-8')
        except UnicodeDecodeError:
            conteudo_texto = conteudo_corrigido_bytes.decode('latin-1', errors='ignore')
    else:
        conteudo_texto = str(conteudo_corrigido_bytes)

    # Identifica se é fatura de cartão de crédito
    is_cartao = bool(re.search(r"<CCSTMTRS>|<CCACCTFROM>", conteudo_texto, re.IGNORECASE))

    # 3. Extração dos dados globais do banco (BANKID / FID)
    bankid_match = re.search(r"<BANKID>\s*([^\r\n<]+)", conteudo_texto, re.IGNORECASE)
    if not bankid_match:
        bankid_match = re.search(r"<FID>\s*([^\r\n<]+)", conteudo_texto, re.IGNORECASE)

    if not bankid_match or not bankid_match.group(1).strip():
        raise HTTPException(
            status_code=400, 
            detail="Arquivo OFX inválido: Tag <BANKID> ou <FID> não encontrada."
        )

    banco_codigo_raw = bankid_match.group(1).strip()
    banco_codigo_limpo = banco_codigo_raw.lstrip("0")

    # Validação do Banco no sistema
    banco_val = mapa_bancos_db.get(banco_codigo_limpo) or mapa_bancos_db.get(banco_codigo_limpo.zfill(3))

    if not banco_val:
        raise HTTPException(
            status_code=400,
            detail=f"O banco com código '{banco_codigo_raw}' informado no OFX não está cadastrado no sistema."
        )

    # Extração de Agência e Conta
    agencia_match = re.search(r"<BRANCHID>\s*([^\r\n<]+)", conteudo_texto, re.IGNORECASE)
    conta_match = re.search(r"<ACCTID>\s*([^\r\n<]+)", conteudo_texto, re.IGNORECASE)

    agencia_val = agencia_match.group(1).strip() if agencia_match else ("CARTAO" if is_cartao else "")
    conta_val = conta_match.group(1).strip() if conta_match else ""

    if conta_val:
        m_conta_x = re.match(r"^(\d+)-?[xX]$", conta_val.strip())
        if m_conta_x:
            conta_val = f"{m_conta_x.group(1)}-0"
        else:
            conta_val = re.sub(r"[xX]$", "0", conta_val.strip())

    # 4. Leitura dos blocos de transação (STMTTRN ou TRN)
    blocos_transacao = re.findall(r"<STMTTRN>(.*?)</STMTTRN>", conteudo_texto, re.DOTALL | re.IGNORECASE)
    if not blocos_transacao:
        blocos_transacao = re.findall(r"<TRN>(.*?)</TRN>", conteudo_texto, re.DOTALL | re.IGNORECASE)

    if not blocos_transacao and "<STMTTRN>" in conteudo_texto.upper():
        partes = re.split(r'<STMTTRN>', conteudo_texto, flags=re.IGNORECASE)
        blocos_transacao = partes[1:]

    transacoes_dados = []

    for bloco in blocos_transacao:
        def extrair_tag(tag_nome, texto_bloco):
            m = re.search(rf'<{tag_nome}>\s*([^<\r\n]+)(?:</{tag_nome}>)?', texto_bloco, re.IGNORECASE)
            return m.group(1).strip() if m else ""

        tipo_bruto = extrair_tag("TRNTYPE", bloco).upper()
        data_str = extrair_tag("DTPOSTED", bloco)
        valor_str = extrair_tag("TRNAMT", bloco)
        memo_raw = extrair_tag("MEMO", bloco)
        payee_raw = extrair_tag("NAME", bloco)
        checknum_raw = extrair_tag("CHECKNUM", bloco)

        # Formatação de Data
        advert_limpa_data = re.sub(r'\D', '', data_str)
        if len(advert_limpa_data) >= 8:
            ano, mes, dia = advert_limpa_data[0:4], advert_limpa_data[4:6], advert_limpa_data[6:8]
            data_formatada = f"{dia}/{mes}/{ano}"
        else:
            data_formatada = data_str[:10]

        # Tratamento de Valor
        valor_limpo = valor_str.replace(",", ".").replace("R$", "").strip()
        try:
            valor = float(valor_limpo)
        except ValueError:
            valor = 0.0

        memo = corrigir_encoding(memo_raw) if 'corrigir_encoding' in globals() else memo_raw
        payee = corrigir_encoding(payee_raw) if 'corrigir_encoding' in globals() else payee_raw
        checknum_val = corrigir_encoding(checknum_raw) if 'corrigir_encoding' in globals() else checknum_raw

        if payee and memo:
            descricao_original = f"{payee} - {memo}"
        else:
            descricao_original = payee or memo or "Transação OFX"

        # Determina o TIPO PRIMEIRO
        desc_lower = descricao_original.lower()
        if "saldo" in desc_lower:
            tipo = "SALDO"
        elif tipo_bruto in ["DEBIT", "DEB"] or valor < 0:
            tipo = "PAGAMENTO"
        else:
            tipo = "RECEBIMENTO"

        # Identifica o fornecedor
        fornecedor_raw = identificar_fornecedor(
            descricao_original, 
            banco_val, 
            tipo=tipo, 
            conexao=conexao
        )

        if isinstance(fornecedor_raw, dict):
            fornecedor_val = (
                fornecedor_raw.get("fornecedor") 
                or fornecedor_raw.get("nome") 
                or ""
            )
        else:
            fornecedor_val = str(fornecedor_raw) if fornecedor_raw is not None else ""

        if "saldo" in fornecedor_val.lower():
            tipo = "SALDO"

        transacoes_dados.append({
            "contratante": "",
            "unidade": "",
            "banco": banco_val,
            "agencia": agencia_val,
            "conta": conta_val,
            "data": data_formatada,
            "descricao": descricao_original,
            "obs": checknum_val,
            "valor": float(valor),
            "tipo": tipo,
            "fornecedor": fornecedor_val,
            "cpf_cnpj": "",
            "planoConta": "",
            "grupoConta": "",
            "edre": ""
        })

    # 5. Consulta e vinculação de Unidade e Contratante (JOIN corrigido)
    if banco_codigo_limpo and conta_val:
        try:
            # Novo JOIN: BancoConta aponta para Unidade (bc.unidadeId = u.id)
            sql_unidade = """
                SELECT TOP 1 u.nome AS unidade_nome, c.nome AS contratante_nome
                FROM dbo.BancoConta bc
                INNER JOIN dbo.Banco b ON bc.bancoId = b.id
                INNER JOIN dbo.Unidade u ON bc.unidadeId = u.id
                LEFT JOIN dbo.Contratante c ON u.contratanteId = c.id
                WHERE (
                    LTRIM(b.codigo, '0') = ? 
                    OR b.codigo = ?
                )
                AND LTRIM(RTRIM(bc.conta)) = ?
            """
            params = [banco_codigo_limpo, banco_codigo_raw, conta_val]

            # Se não for cartão de crédito e houver agência, adiciona a validação da agência na query
            if not is_cartao and agencia_val:
                sql_unidade += " AND LTRIM(RTRIM(bc.agencia)) = ?"
                params.append(agencia_val)

            cursor.execute(sql_unidade, params)
            row_unidade = cursor.fetchone()

            if row_unidade:
                unidade_encontrada = row_unidade[0] or ""
                contratante_encontrado = row_unidade[1] or ""
                
                # Aplica Unidade e Contratante para todas as transações extraídas
                for t in transacoes_dados:
                    t["unidade"] = unidade_encontrada
                    t["contratante"] = contratante_encontrado
        except Exception:
            pass  # Mantém os valores vazios em caso de erro

    return transacoes_dados

def processar_pdf(conteudo_bytes: bytes, conexao) -> list:
    banco_detectado = identificar_banco_no_pdf(conteudo_bytes).strip()

    if not banco_detectado:
        raise HTTPException(
            status_code=400, 
            detail="Não foi possível identificar o banco emissor deste PDF."
        )

    # Wrapper que aceita o parâmetro 'tipo' vindo das funções de PDF
    def identificar_fn(descricao: str, banco_v: str, tipo: str = ""):
        # Caso a função de leitura do PDF não passe o tipo, deduz a partir da descrição
        if not tipo:
            desc_lower = (descricao or "").lower()
            if "saldo" in desc_lower:
                tipo = "SALDO"

        return identificar_fornecedor(descricao, banco_v, tipo=tipo, conexao=conexao)

    if banco_detectado == "BANCO DO BRASIL":
        dados_extraidos = pdf_banco_do_brasil(conteudo_bytes, identificar_fn)
    elif banco_detectado == "V&P":
        dados_extraidos = pdf_vp(conteudo_bytes, identificar_fn)
    elif banco_detectado == "BRADESCO":
        dados_extraidos = pdf_bradesco(conteudo_bytes, identificar_fn)
    elif banco_detectado in ["CAIXA ECONOMICA FEDERAL", "CAIXA"]:
        dados_extraidos = pdf_caixa(conteudo_bytes, identificar_fn)
    elif banco_detectado == "SICOOB":
        dados_extraidos = pdf_sicoob(conteudo_bytes, identificar_fn)
    elif banco_detectado == "BANCO SAFRA":
        dados_extraidos = pdf_safra(conteudo_bytes, identificar_fn)
    elif banco_detectado == "SICREDI":
        dados_extraidos = pdf_sicredi(conteudo_bytes, identificar_fn)
    elif banco_detectado == "BANCO C6":
        dados_extraidos = pdf_c6(conteudo_bytes, identificar_fn)
    elif banco_detectado == "BANCO INTER":
        dados_extraidos = pdf_inter(conteudo_bytes, identificar_fn)
    elif banco_detectado == "BANCO CORA":
        dados_extraidos = pdf_cora(conteudo_bytes, identificar_fn)
    elif banco_detectado == "BANCO STONE":
        dados_extraidos = pdf_stone(conteudo_bytes, identificar_fn)
    else:
        raise HTTPException(
            status_code=400, 
            detail=f"O processamento do PDF para o banco '{banco_detectado}' ainda não foi implementado."
        )

    transacoes = dados_extraidos.get("transacoes", [])
    agencia_val = dados_extraidos.get("agencia", "")
    conta_val = dados_extraidos.get("conta", "")

    # Re-aplica a identificação do Fornecedor caso o PDF tenha montado o TIPO depois de extrair os valores
    for t in transacoes:
        tipo_transacao = t.get("tipo", "")
        # Se o fornecedor veio vazio e o tipo já foi identificado pelo parser do PDF (PAGAMENTO / RECEBIMENTO)
        if not t.get("fornecedor") and tipo_transacao:
            t["fornecedor"] = identificar_fornecedor(
                t.get("descricao", ""), 
                banco_detectado, 
                tipo=tipo_transacao, 
                conexao=conexao
            )

    # Busca Unidade e Contratante no Banco de Dados
    unidade_val = ""
    contratante_val = ""

    if conexao and agencia_val and conta_val:
        cursor = conexao.cursor()
        agencia_limpa = re.sub(r"\D", "", agencia_val).lstrip("0")
        conta_limpa = re.sub(r"\D", "", conta_val).lstrip("0")

        # Query atualizada: BancoConta possui a FK unidadeId
        cursor.execute(
            """
            SELECT u.nome AS unidade_nome, c.nome AS contratante_nome
            FROM dbo.BancoConta bc
            INNER JOIN dbo.Banco b ON bc.bancoId = b.id
            LEFT JOIN dbo.Unidade u ON bc.unidadeId = u.id
            LEFT JOIN dbo.Contratante c ON u.contratanteId = c.id
            WHERE LTRIM(RTRIM(REPLACE(bc.agencia, '-', ''))) LIKE ?
              AND LTRIM(RTRIM(REPLACE(bc.conta, '-', ''))) LIKE ?
            """,
            (f"%{agencia_limpa}%", f"%{conta_limpa}%")
        )
        row_unidade = cursor.fetchone()
        if row_unidade:
            unidade_val = row_unidade[0] or ""
            contratante_val = row_unidade[1] or ""

    # Aplica Unidade e Contratante nas transações
    for t in transacoes:
        t["contratante"] = contratante_val
        t["unidade"] = unidade_val or t.get("unidade", "")

    return transacoes

def processar_arquivo(file: UploadFile, conteudo_bytes: bytes) -> list:
    filename_lower = file.filename.lower()

    if filename_lower.endswith(".ofx"):
        conteudo_texto = None
        for encoding in ["cp1252", "latin-1", "utf-8"]:
            try:
                conteudo_texto = conteudo_bytes.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if not conteudo_texto:
            raise HTTPException(status_code=400, detail="Não foi possível decodificar o arquivo OFX.")
        
        conexao = obter_conexao(BANCO_AUTENTICACAO)
        try:
            return processar_ofx(conteudo_texto, conexao)
        finally:
            conexao.close()

    elif filename_lower.endswith(".pdf"):
        conexao = obter_conexao(BANCO_AUTENTICACAO)
        try:
            return processar_pdf(conteudo_bytes, conexao)
        finally:
            conexao.close()

    else:
        raise HTTPException(
            status_code=400, 
            detail="Formato de arquivo não suportado. Envie um arquivo .ofx ou .pdf"
        )

# ==============================================================================
# EXCEL FORMULATION
# ==============================================================================

def padronizar_dataframe(transacoes: list) -> pd.DataFrame:
    df = pd.DataFrame(transacoes)

    if "valor" in df.columns:
        df["valor"] = pd.to_numeric(df["valor"], errors="coerce").fillna(0.0)

    # Renomeia colunas do dicionário Python para o cabeçalho final do Excel
    mapeamento_colunas = {
        "contratante": "CONTRATANTE",
        "unidade": "UNIDADE",
        "banco": "BANCO",
        "agencia": "AGÊNCIA",
        "conta": "CONTA",
        "data": "DATA",
        "descricao": "DESCRIÇÃO",
        "obs": "OBSERVAÇÃO",
        "valor": "VALOR",
        "tipo": "TIPO",
        "fornecedor": "FORNECEDOR",
        "cpf_cnpj": "CPF_CNPJ",
        "planoConta": "PLANO DE CONTA",
        "grupoConta": "GRUPO DE CONTA",
        "edre": "E-DRE"
    }

    df = df.rename(columns=mapeamento_colunas)

    ordem_desejada = [
        "CONTRATANTE", "UNIDADE", "BANCO", "AGÊNCIA", "CONTA",
        "DATA", "DESCRIÇÃO", "OBSERVAÇÃO", "VALOR", "TIPO", "FORNECEDOR",
        "CPF_CNPJ", "PLANO DE CONTA", "GRUPO DE CONTA", "E-DRE"
    ]

    return df[[col for col in ordem_desejada if col in df.columns]]

# ==============================================================================
# ENDPOINTS DA API
# ==============================================================================

@router.post("/preview")
async def converter_preview(
    file: UploadFile = File(...),
    contratanteId: Optional[int] = Form(None),
):
    conteudo_bytes = await file.read()
    # Passa o 'banco' aqui para abrir a conexão
    transacoes = processar_arquivo(file, conteudo_bytes) 
    
    transacoes = aplicar_plano_conta(transacoes, contratante_id=contratanteId)
    return {"transacoes": transacoes}


@router.post("/download")
async def converter_download(
    file: UploadFile = File(...),
    contratanteId: Optional[int] = Form(None),
):
    try:
        conteudo_bytes = await file.read()
        transacoes = processar_arquivo(file, conteudo_bytes)

        if not transacoes:
            raise HTTPException(status_code=400, detail="Nenhuma transação encontrada no arquivo.")

        # Aplica as regras no conjunto de dados antes de gerar a planilha Excel
        transacoes = aplicar_plano_conta(transacoes, contratante_id=contratanteId)

        df = padronizar_dataframe(transacoes)
        nome_aba = "BASE_FINANCEIRA"
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=nome_aba)
            worksheet = writer.sheets[nome_aba]

            fill_azul = PatternFill(start_color="35448A", end_color="35448A", fill_type="solid")
            fonte_branca = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            alinhamento_centro = Alignment(horizontal="center", vertical="center")

            for cell in worksheet[1]:
                cell.fill = fill_azul
                cell.font = fonte_branca
                cell.alignment = alinhamento_centro

            if "VALOR" in df.columns:
                col_valor_idx = df.columns.get_loc("VALOR") + 1
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=col_valor_idx)
                    cell.value = float(cell.value or 0.0)
                    cell.number_format = "R$ #,##0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="center")

            for col in worksheet.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                col_letter = col[0].column_letter
                worksheet.column_dimensions[col_letter].width = max(max_len + 10, 12)

        output.seek(0)
        nome_arquivo_saida = f"{file.filename.lower().rsplit('.', 1)[0]}_convertido.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={nome_arquivo_saida}"},
        )
        
    except HTTPException as he:
        raise he
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erro interno ao gerar Excel: {str(e)}")